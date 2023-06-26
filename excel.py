import dataclasses
import xlsxwriter
from openpyxl import load_workbook
import helper


@dataclasses.dataclass
class Rows:
    day: str
    league: str
    home: str
    defence: str


old_m = []
new_m = []

filename_new = helper.vals['filename_new']
filename_old = helper.vals['filename_old']

workbook_new = load_workbook(filename=filename_new)
workbook_old = load_workbook(filename=filename_old)
sheet_new = workbook_new.active
sheet_old = workbook_old.active


def excel_function(list_match, sheet_):
    m_row = sheet_.max_row
    for i in range(1, m_row + 1):
        row_ = Rows(day=sheet_.cell(row=i, column=1).value,
                    league=sheet_.cell(row=i, column=2).value,
                    home=sheet_.cell(row=i, column=3).value, defence=sheet_.cell(row=i, column=4).value)
        list_match.append(row_)
    return list_match


new_m = excel_function(new_m, sheet_new)
old_m = excel_function(old_m, sheet_old)

# print(new_m[0])
# print(old_m[0])

m_col = sheet_new.max_column

output_file = helper.vals['output_file']
workbook = xlsxwriter.Workbook(output_file)
sheet = workbook.add_worksheet()

week_val = helper.vals['week_val']

row_len_old = len(old_m)
row_len_new = len(new_m)
row_count = 0

for index_1 in range(0, row_len_old):
    for index_2 in range(0, row_len_new):
        if old_m[index_1].week_one == new_m[index_2].week_one and old_m[index_1].league == new_m[index_2].league:
            if old_m[index_1].home == new_m[index_2].home and old_m[index_1].defence == new_m[index_2].defence:
                for index_col in range(1, m_col + 1):
                    if index_col == 5 or index_col == 6:
                        sheet.write(row_count, index_col - 1, sheet_new.cell(row=index_2 + 1, column=index_col).value)
                    else:
                        sheet.write(row_count, index_col - 1, sheet_old.cell(row=index_1 + 1, column=index_col).value)

                row_count = row_count + 1
                break

workbook.close()

